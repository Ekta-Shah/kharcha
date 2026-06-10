import io
from collections import defaultdict
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import BankTransaction, Expense, Match

_INK = "173F35"
_RED = "AE2B26"
_PAPER = "F7F6F2"
_HEADER_FILL = PatternFill("solid", fgColor=_INK)
_TOTAL_FILL = PatternFill("solid", fgColor="E8F0EE")
_DOUBLE_BOTTOM = Border(bottom=Side(style="double", color=_RED))
_BOLD = Font(bold=True)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_RED_BOLD = Font(bold=True, color=_RED)
_MONO = "Courier New"


def _hdr(ws, col: int, text: str) -> None:
    c = ws.cell(row=1, column=col, value=text)
    c.fill = _HEADER_FILL
    c.font = _HEADER_FONT
    c.alignment = Alignment(horizontal="center")


async def export(from_date: date | None, to_date: date | None, db: AsyncSession) -> io.BytesIO:
    q = select(Expense).order_by(Expense.date.asc(), Expense.created_at.asc())
    if from_date:
        q = q.where(Expense.date >= from_date)
    if to_date:
        q = q.where(Expense.date <= to_date)
    result = await db.execute(q)
    expenses = result.scalars().all()

    wb = Workbook()

    # Sheet 1: Line Items
    ws1 = wb.active
    ws1.title = "Line Items"
    for i, hdr in enumerate(["Date", "Item", "Category", "Amount (₹)", "Source"], 1):
        _hdr(ws1, i, hdr)
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 10

    by_date: dict[date, list[Expense]] = defaultdict(list)
    for exp in expenses:
        by_date[exp.date].append(exp)

    row = 2
    for day in sorted(by_date):
        day_exps = by_date[day]
        for exp in day_exps:
            ws1.cell(row=row, column=1, value=exp.date.isoformat())
            ws1.cell(row=row, column=2, value=exp.item)
            ws1.cell(row=row, column=3, value=exp.category)
            amt = ws1.cell(row=row, column=4, value=float(exp.cost))
            amt.number_format = '#,##0.00'
            amt.font = Font(name=_MONO)
            ws1.cell(row=row, column=5, value=exp.source)
            row += 1

        # daily total row
        day_total = sum(e.cost for e in day_exps)
        ws1.cell(row=row, column=1, value=day.isoformat()).font = _RED_BOLD
        ws1.cell(row=row, column=2, value=f"Daily Total ({len(day_exps)} items)").font = _RED_BOLD
        tot = ws1.cell(row=row, column=4, value=float(day_total))
        tot.font = _RED_BOLD
        tot.number_format = '#,##0.00'
        for col in range(1, 6):
            ws1.cell(row=row, column=col).border = _DOUBLE_BOTTOM
            ws1.cell(row=row, column=col).fill = _TOTAL_FILL
        row += 1

    # Sheet 2: Daily Totals
    ws2 = wb.create_sheet("Daily Totals")
    for i, hdr in enumerate(["Date", "Items", "Total (₹)"], 1):
        _hdr(ws2, i, hdr)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 14

    grand = Decimal(0)
    for r, day in enumerate(sorted(by_date), start=2):
        day_exps = by_date[day]
        day_total = sum(e.cost for e in day_exps)
        grand += day_total
        ws2.cell(row=r, column=1, value=day.isoformat())
        ws2.cell(row=r, column=2, value=len(day_exps))
        tot = ws2.cell(row=r, column=3, value=float(day_total))
        tot.number_format = '#,##0.00'
        tot.font = Font(name=_MONO)

    last = len(by_date) + 2
    ws2.cell(row=last, column=1, value="Grand Total").font = _RED_BOLD
    gt = ws2.cell(row=last, column=3, value=float(grand))
    gt.number_format = '#,##0.00'
    gt.font = _RED_BOLD
    for col in range(1, 4):
        ws2.cell(row=last, column=col).border = _DOUBLE_BOTTOM
        ws2.cell(row=last, column=col).fill = _TOTAL_FILL

    # Sheet 3: Reconciliation
    ws3 = wb.create_sheet("Reconciliation")
    recon_hdrs = ["Date", "Ledger Item", "Category", "Ledger ₹", "Bank Description", "Bank ₹", "Status", "Confidence"]
    for i, h in enumerate(recon_hdrs, 1):
        _hdr(ws3, i, h)
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 36
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 10
    ws3.column_dimensions["H"].width = 12

    matches_res = await db.execute(select(Match))
    all_matches = matches_res.scalars().all()

    exp_ids = [m.expense_id for m in all_matches if m.expense_id]
    txn_ids = [m.bank_txn_id for m in all_matches if m.bank_txn_id]
    exp_map: dict = {}
    txn_map: dict = {}
    if exp_ids:
        er = await db.execute(select(Expense).where(Expense.id.in_(exp_ids)))
        exp_map = {e.id: e for e in er.scalars().all()}
    if txn_ids:
        tr = await db.execute(select(BankTransaction).where(BankTransaction.id.in_(txn_ids)))
        txn_map = {t.id: t for t in tr.scalars().all()}

    for r, m in enumerate(all_matches, start=2):
        exp = exp_map.get(m.expense_id)
        txn = txn_map.get(m.bank_txn_id)
        ws3.cell(row=r, column=1, value=exp.date.isoformat() if exp else (txn.txn_date.isoformat() if txn else ""))
        ws3.cell(row=r, column=2, value=exp.item if exp else "")
        ws3.cell(row=r, column=3, value=exp.category if exp else "")
        c4 = ws3.cell(row=r, column=4, value=float(exp.cost) if exp else "")
        if exp:
            c4.number_format = '#,##0.00'
            c4.font = Font(name=_MONO)
        ws3.cell(row=r, column=5, value=txn.description if txn else "")
        c6 = ws3.cell(row=r, column=6, value=float(txn.amount) if txn else "")
        if txn:
            c6.number_format = '#,##0.00'
            c6.font = Font(name=_MONO)
        ws3.cell(row=r, column=7, value=m.status)
        ws3.cell(row=r, column=8, value=float(m.confidence) if m.confidence else "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
